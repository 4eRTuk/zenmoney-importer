import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from adapters.base import BaseAdapter
from models import TransactionRecord
from settings import SourceConfig

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = frozenset({"Date", "Direction", "Comment", "Amount", "Currency"})
SUPPORTED_DIRECTIONS = frozenset({"Coupon", "Dividends", "Taxes"})
TICKER_PARENS_PATTERN = re.compile(r"\(([^()]*)\)")


def parse_tradernet_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            raise ValueError("Date is empty")

        for date_format in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%m/%d/%y",
            "%m/%d/%Y",
            "%d.%m.%Y",
        ):
            try:
                return datetime.strptime(normalized, date_format).date()
            except ValueError:
                continue

        try:
            return datetime.fromisoformat(normalized).date()
        except ValueError as exc:
            raise ValueError(f"Unsupported date format: {value!r}") from exc

    raise TypeError(f"Unsupported date type: {type(value)!r}")


def parse_tradernet_amount(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        normalized = value.strip().replace(",", "")
        if not normalized:
            raise ValueError("Amount is empty")
        return Decimal(normalized)
    raise TypeError(f"Unsupported amount type: {type(value)!r}")


def extract_tradernet_ticker(comment: str) -> str | None:
    if not comment:
        return None

    matches = TICKER_PARENS_PATTERN.findall(comment)
    if not matches:
        return None

    ticker = matches[-1].strip()
    return ticker or None


class TradernetAdapter(BaseAdapter):
    def __init__(
        self,
        source_config: SourceConfig,
        income_category: str,
        adjustment_category: str,
    ) -> None:
        self.default_account = source_config.default_account
        self.income_category = income_category
        self.adjustment_category = adjustment_category

    def parse(self, file_path: Path) -> list[TransactionRecord]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)

            header_row = next(rows_iter, None)
            if header_row is None:
                raise ValueError(f"No data in file: {file_path.name}")

            header_map = {
                str(column).strip(): index
                for index, column in enumerate(header_row)
                if column is not None
            }

            missing_columns = REQUIRED_COLUMNS - set(header_map.keys())
            if missing_columns:
                raise ValueError(
                    f"Missing required columns in {file_path.name}: "
                    f"{', '.join(sorted(missing_columns))}"
                )

            records: list[TransactionRecord] = []
            row_number = 1

            for row in rows_iter:
                row_number += 1
                if row is None or all(cell is None for cell in row):
                    continue

                currency_value = row[header_map["Currency"]]
                currency = str(currency_value).strip() if currency_value is not None else ""
                if currency != "USD":
                    continue

                direction_value = row[header_map["Direction"]]
                direction = (
                    str(direction_value).strip() if direction_value is not None else ""
                )
                if direction not in SUPPORTED_DIRECTIONS:
                    continue

                date_value = row[header_map["Date"]]
                if date_value is None or (
                    isinstance(date_value, str) and not date_value.strip()
                ):
                    logger.warning("Skipping row %d: empty Date", row_number)
                    continue

                try:
                    parsed_date = parse_tradernet_date(date_value)
                except (ValueError, TypeError) as exc:
                    logger.warning(
                        "Skipping row %d: invalid Date %r: %s",
                        row_number,
                        date_value,
                        exc,
                    )
                    continue

                amount_value = row[header_map["Amount"]]
                if amount_value is None or (
                    isinstance(amount_value, str) and not amount_value.strip()
                ):
                    logger.warning("Skipping row %d: empty Amount", row_number)
                    continue

                try:
                    amount = parse_tradernet_amount(amount_value)
                except (InvalidOperation, ValueError, TypeError) as exc:
                    logger.warning(
                        "Skipping row %d: invalid Amount %r: %s",
                        row_number,
                        amount_value,
                        exc,
                    )
                    continue

                comment_value = row[header_map["Comment"]]
                comment_text = (
                    str(comment_value).strip() if comment_value is not None else ""
                )
                ticker = extract_tradernet_ticker(comment_text)

                if direction == "Taxes":
                    category = self.adjustment_category
                    income_amount = None
                    expense_amount = abs(amount)
                else:
                    category = self.income_category
                    income_amount = amount
                    expense_amount = None

                records.append(
                    TransactionRecord(
                        date=parsed_date,
                        category=category,
                        account=self.default_account,
                        income_amount=income_amount,
                        expense_amount=expense_amount,
                        comment=ticker or "",
                        source_type="tradernet",
                        source_file=file_path.name,
                        row_number=row_number,
                    )
                )
        finally:
            wb.close()

        logger.info(
            "Parsed %d tradernet record(s) from %s",
            len(records),
            file_path.name,
        )
        return records
