import logging
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from adapters.base import BaseAdapter
from dates import last_day_of_month
from models import TransactionRecord
from settings import SourceConfig

logger = logging.getLogger(__name__)

SHEET_NAME = "Отчет о выплате доходов по ЦБ"
REQUIRED_COLUMNS = frozenset(
    {
        "Эмитент ЦБ",
        "Вид ЦБ",
        "Дата выплаты",
        "Начислено (в рублях)",
        "Налог к удержанию",
    }
)
SUPPORTED_SECURITY_TYPES = frozenset({"Акция", "Облигация"})
HEADER_ALIASES = {
    "Эмитент ЦБ": ("эмитентцб",),
    "Вид ЦБ": ("видцб",),
    "Дата выплаты": ("датавыплаты",),
    "Начислено (в рублях)": ("начисленоврублях",),
    "Налог к удержанию": ("налогкудержанию",),
}
FALLBACK_COLUMN_INDEXES = {
    "Эмитент ЦБ": 0,
    "Вид ЦБ": 4,
    "Дата выплаты": 5,
    "Начислено (в рублях)": 9,
    "Налог к удержанию": 16,
}


def normalize_cell_text(value: object) -> str:
    return " ".join(str(value).split())


def normalize_header_key(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", normalize_cell_text(value).casefold())
    return "".join(ch for ch in normalized if ch.isalnum())


def build_header_map(row: tuple[object, ...]) -> dict[str, int]:
    required_keys = {normalize_header_key(name): name for name in REQUIRED_COLUMNS}
    alias_keys = {
        normalize_header_key(alias): canonical_name
        for canonical_name, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }

    normalized_cells: list[tuple[int, str]] = [
        (index, normalize_header_key(cell))
        for index, cell in enumerate(row)
        if cell is not None and normalize_header_key(cell)
    ]

    header_map: dict[str, int] = {}
    for index, normalized_cell in normalized_cells:
        canonical_name = required_keys.get(normalized_cell) or alias_keys.get(normalized_cell)
        if canonical_name:
            header_map[canonical_name] = index

    for canonical_name, aliases in HEADER_ALIASES.items():
        for index, normalized_cell in normalized_cells:
            if normalized_cell in aliases:
                header_map[canonical_name] = index
                break
        else:
            for index in range(len(normalized_cells) - 1):
                combined = normalized_cells[index][1] + normalized_cells[index + 1][1]
                if combined in aliases:
                    header_map[canonical_name] = normalized_cells[index][0]
                    break

    for normalized_name, canonical_name in required_keys.items():
        if canonical_name not in header_map:
            for index, normalized_cell in normalized_cells:
                if normalized_cell == normalized_name:
                    header_map[canonical_name] = index
                    break

    return header_map


def resolve_column_indexes(
    ws,
    file_name: str,
) -> dict[str, int]:
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row is None or all(cell is None for cell in row):
            continue

        candidate_header_map = build_header_map(row)
        if REQUIRED_COLUMNS <= set(candidate_header_map.keys()):
            return candidate_header_map

    logger.warning(
        "Could not detect Tinkoff headers in %s, using fallback column positions",
        file_name,
    )
    return dict(FALLBACK_COLUMN_INDEXES)


def parse_tinkoff_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            raise ValueError("Date is empty")

        for date_format in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(normalized, date_format).date()
            except ValueError:
                continue

        try:
            return datetime.fromisoformat(normalized).date()
        except ValueError as exc:
            raise ValueError(f"Unsupported date format: {value!r}") from exc

    raise TypeError(f"Unsupported date type: {type(value)!r}")


def parse_tinkoff_amount(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        normalized = value.strip().replace(" ", "").replace("\xa0", "")
        if not normalized:
            raise ValueError("Amount is empty")
        normalized = normalized.replace(",", ".")
        return Decimal(normalized)

    raise TypeError(f"Unsupported amount type: {type(value)!r}")


class TinkoffBrokerAdapter(BaseAdapter):
    def __init__(
        self,
        source_config: SourceConfig,
        income_category: str,
        adjustment_category: str,
        target_month: int,
        target_year: int,
    ) -> None:
        if not source_config.default_account:
            raise ValueError("tinkoff_broker source must define default_account")

        self.default_account = source_config.default_account
        self.income_category = income_category
        self.adjustment_category = adjustment_category
        self.target_month = target_month
        self.target_year = target_year
        self.adjustment_date = last_day_of_month(target_month, target_year)

    def parse(self, file_path: Path) -> list[TransactionRecord]:
        wb = load_workbook(file_path, read_only=False, data_only=True)
        try:
            if SHEET_NAME not in wb.sheetnames:
                raise ValueError(
                    f"Missing required sheet in {file_path.name}: {SHEET_NAME}"
                )

            ws = wb[SHEET_NAME]
            header_map = resolve_column_indexes(ws, file_path.name)

            records: list[TransactionRecord] = []
            tax_total = Decimal("0")

            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row is None or all(cell is None for cell in row):
                    continue

                if len(row) <= max(header_map.values()):
                    continue

                security_type_value = row[header_map["Вид ЦБ"]]
                security_type = (
                    normalize_cell_text(security_type_value)
                    if security_type_value is not None
                    else ""
                )
                if security_type not in SUPPORTED_SECURITY_TYPES:
                    continue

                date_value = row[header_map["Дата выплаты"]]
                if date_value is None or (
                    isinstance(date_value, str) and not date_value.strip()
                ):
                    logger.warning("Skipping row %d: empty Дата выплаты", row_number)
                    continue

                try:
                    parsed_date = parse_tinkoff_date(date_value)
                except (ValueError, TypeError) as exc:
                    logger.warning(
                        "Skipping row %d: invalid Дата выплаты %r: %s",
                        row_number,
                        date_value,
                        exc,
                    )
                    continue

                if (
                    parsed_date.month != self.target_month
                    or parsed_date.year != self.target_year
                ):
                    continue

                amount_value = row[header_map["Начислено (в рублях)"]]
                if amount_value is None or (
                    isinstance(amount_value, str) and not amount_value.strip()
                ):
                    logger.warning("Skipping row %d: empty Начислено (в рублях)", row_number)
                    continue

                try:
                    income_amount = parse_tinkoff_amount(amount_value)
                except (InvalidOperation, ValueError, TypeError) as exc:
                    logger.warning(
                        "Skipping row %d: invalid Начислено (в рублях) %r: %s",
                        row_number,
                        amount_value,
                        exc,
                    )
                    continue

                issuer_value = row[header_map["Эмитент ЦБ"]]
                comment = (
                    normalize_cell_text(issuer_value) if issuer_value is not None else ""
                )

                records.append(
                    TransactionRecord(
                        date=parsed_date,
                        category=self.income_category,
                        account=self.default_account,
                        income_amount=income_amount,
                        expense_amount=None,
                        comment=comment,
                        source_type="tinkoff_broker",
                        source_file=file_path.name,
                        row_number=row_number,
                    )
                )

                if security_type != "Акция":
                    continue

                tax_value = row[header_map["Налог к удержанию"]]
                if tax_value is None or (
                    isinstance(tax_value, str) and not tax_value.strip()
                ):
                    logger.warning("Skipping tax on row %d: empty Налог к удержанию", row_number)
                    continue

                try:
                    tax_total += parse_tinkoff_amount(tax_value)
                except (InvalidOperation, ValueError, TypeError) as exc:
                    logger.warning(
                        "Skipping tax on row %d: invalid Налог к удержанию %r: %s",
                        row_number,
                        tax_value,
                        exc,
                    )

            records.append(
                TransactionRecord(
                    date=self.adjustment_date,
                    category=self.adjustment_category,
                    account=self.default_account,
                    income_amount=None,
                    expense_amount=tax_total,
                    comment="",
                    source_type="tinkoff_broker",
                    source_file=file_path.name,
                )
            )
        finally:
            wb.close()

        logger.info(
            "Parsed %d tinkoff_broker record(s) from %s",
            len(records),
            file_path.name,
        )
        return records
