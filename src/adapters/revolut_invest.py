import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from adapters.base import BaseAdapter
from models import TransactionRecord
from settings import SourceConfig

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = frozenset({"Date", "Ticker", "Type", "Total Amount"})
MONEY_AMOUNT_PATTERN = re.compile(r"^USD\s+(.+)$", re.IGNORECASE)


def parse_revolut_datetime(value: str) -> date:
    normalized = value.strip()
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"
    return datetime.fromisoformat(normalized).date()


def parse_money_amount(value: str) -> Decimal:
    match = MONEY_AMOUNT_PATTERN.match(value.strip())
    if not match:
        raise ValueError(f"Unsupported amount format: {value!r}")
    return Decimal(match.group(1).strip())


class RevolutInvestAdapter(BaseAdapter):
    def __init__(self, source_config: SourceConfig, income_category: str) -> None:
        self.default_account = source_config.default_account
        self.income_category = income_category

    def parse(self, file_path: Path) -> list[TransactionRecord]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)

            header_row = next(rows_iter, None)
            if header_row is None:
                raise ValueError(f"No data in file: {file_path.name}")

            header_map = {
                str(column).strip(): index
                for index, column in enumerate(header_row)
                if column is not None
            }

            missing_columns = REQUIRED_COLUMNS - set(header_map.keys())
            if missing_columns:
                raise ValueError(
                    f"Missing required columns in {file_path.name}: "
                    f"{', '.join(sorted(missing_columns))}"
                )

            records: list[TransactionRecord] = []
            row_number = 1

            for row in rows_iter:
                row_number += 1
                if row is None or all(cell is None for cell in row):
                    continue

                row_type = row[header_map["Type"]]
                if row_type is None or str(row_type).strip() != "DIVIDEND":
                    continue

                date_value = row[header_map["Date"]]
                if date_value is None or (
                    isinstance(date_value, str) and not date_value.strip()
                ):
                    logger.warning("Skipping row %d: empty Date", row_number)
                    continue

                try:
                    parsed_date = parse_revolut_datetime(str(date_value))
                except (ValueError, TypeError) as exc:
                    logger.warning(
                        "Skipping row %d: invalid Date %r: %s",
                        row_number,
                        date_value,
                        exc,
                    )
                    continue

                ticker = row[header_map["Ticker"]]
                comment = str(ticker).strip() if ticker is not None else ""

                amount_value = row[header_map["Total Amount"]]
                if amount_value is None or (
                    isinstance(amount_value, str) and not amount_value.strip()
                ):
                    logger.warning("Skipping row %d: empty Total Amount", row_number)
                    continue

                try:
                    income_amount = parse_money_amount(str(amount_value))
                except (ValueError, InvalidOperation) as exc:
                    logger.warning(
                        "Skipping row %d: invalid Total Amount %r: %s",
                        row_number,
                        amount_value,
                        exc,
                    )
                    continue

                records.append(
                    TransactionRecord(
                        date=parsed_date,
                        category=self.income_category,
                        account=self.default_account,
                        income_amount=income_amount,
                        expense_amount=None,
                        comment=comment,
                        source_type="revolut_invest",
                        source_file=file_path.name,
                        row_number=row_number,
                    )
                )
        finally:
            wb.close()

        logger.info(
            "Parsed %d dividend record(s) from %s",
            len(records),
            file_path.name,
        )
        return records
